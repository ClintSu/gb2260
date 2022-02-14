#!/usr/bin/python
# -*- coding: UTF-8 -*- 

class Area:
    def __init__(self,revision,code,name):
        self.revision = revision
        self.code = code
        self.name = name 

#手动粘贴每年数据到excel中 http://www.mca.gov.cn/article/sj/xzqh/1980/

from math import fabs
from operator import attrgetter
import openpyxl

wb = openpyxl.Workbook()
wb = openpyxl.load_workbook('gb2260.xlsx')

# print(wb.sheetnames)


dict ={}
list =[]
for sheet in wb:
    print(sheet.title)
    ws  = wb[sheet.title]
    for rx in range(1,ws.max_row+1):
        w1 = str(ws.cell(rx,1).value).replace(' ','')
        w2 = str(ws.cell(rx,2).value).replace(' ','')

        dict[w1] = w2
        area = Area(sheet.title,w1,w2)
        list.append(area) 


# 打开一个文件
fo = open("gb2260.sql", "w", encoding="utf-8")

#for key in dict:
     #print(key,dict[key])  
    # fo.write(ww)

list.sort(key= attrgetter('code'))
sortedlist = sorted(list,key = attrgetter('revision'),reverse=False)

for item in sortedlist:
    sw ='insert into gb2260 (revision,code,name) values (\''+ item.revision+'\',\''+ item.code+'\',\''+item.name+ '\');';
    print(sw)
    fo.write(sw+'\n')
# 关闭打开的文件
fo.close()